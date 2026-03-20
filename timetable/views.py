from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db.models import Count
from .models import (Department, Faculty, Classroom, Course, TimeSlot, 
                    StudentGroup, CourseAssignment, TimetableEntry)
from .generator import TimetableGenerator
import io
import csv
from datetime import datetime

def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        user_type = request.POST.get('user_type')
        
        if password1 != password2:
            messages.error(request, 'Passwords do not match')
            return redirect('register')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists')
            return redirect('register')
        
        user = User.objects.create_user(username=username, email=email, password=password1)
        user.is_staff = (user_type == 'admin')
        user.save()
        
        messages.success(request, 'Registration successful! Please login.')
        return redirect('login')
    return render(request, 'register.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    stats = {
        'departments': Department.objects.count(),
        'faculty': Faculty.objects.count(),
        'courses': Course.objects.count(),
        'classrooms': Classroom.objects.count(),
        'time_slots': TimeSlot.objects.count(),
        'student_groups': StudentGroup.objects.count(),
        'assignments': CourseAssignment.objects.count(),
        'entries': TimetableEntry.objects.count(),
    }
    return render(request, 'dashboard.html', {'stats': stats})

@login_required
def department_list(request):
    departments = Department.objects.all().order_by('code')
    return render(request, 'department_list.html', {'departments': departments})

@login_required
def department_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        Department.objects.create(name=name, code=code)
        messages.success(request, 'Department added successfully')
        return redirect('department_list')
    return render(request, 'department_form.html')

@login_required
def department_edit(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        dept.name = request.POST.get('name')
        dept.code = request.POST.get('code')
        dept.save()
        messages.success(request, 'Department updated successfully')
        return redirect('department_list')
    return render(request, 'department_form.html', {'department': dept})

@login_required
def department_delete(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    dept.delete()
    messages.success(request, 'Department deleted successfully')
    return redirect('department_list')

@login_required
def faculty_list(request):
    faculty = Faculty.objects.select_related('department').all()
    return render(request, 'faculty_list.html', {'faculty': faculty})

@login_required
def faculty_add(request):
    departments = Department.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        department_id = request.POST.get('department')
        available_days = request.POST.getlist('available_days')
        
        department = Department.objects.get(pk=department_id)
        Faculty.objects.create(
            name=name, email=email, phone=phone,
            department=department, available_days=available_days
        )
        messages.success(request, 'Faculty added successfully')
        return redirect('faculty_list')
    return render(request, 'faculty_form.html', {'departments': departments})

@login_required
def faculty_edit(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    departments = Department.objects.all()
    if request.method == 'POST':
        faculty.name = request.POST.get('name')
        faculty.email = request.POST.get('email')
        faculty.phone = request.POST.get('phone')
        faculty.department = Department.objects.get(pk=request.POST.get('department'))
        faculty.available_days = request.POST.getlist('available_days')
        faculty.save()
        messages.success(request, 'Faculty updated successfully')
        return redirect('faculty_list')
    return render(request, 'faculty_form.html', {'faculty': faculty, 'departments': departments})

@login_required
def faculty_delete(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    faculty.delete()
    messages.success(request, 'Faculty deleted successfully')
    return redirect('faculty_list')

@login_required
def classroom_list(request):
    classrooms = Classroom.objects.all().order_by('building', 'name')
    return render(request, 'classroom_list.html', {'classrooms': classrooms})

@login_required
def classroom_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        building = request.POST.get('building')
        capacity = request.POST.get('capacity')
        room_type = request.POST.get('room_type')
        Classroom.objects.create(name=name, building=building, 
                                capacity=capacity, room_type=room_type)
        messages.success(request, 'Classroom added successfully')
        return redirect('classroom_list')
    return render(request, 'classroom_form.html')

@login_required
def classroom_edit(request, pk):
    room = get_object_or_404(Classroom, pk=pk)
    if request.method == 'POST':
        room.name = request.POST.get('name')
        room.building = request.POST.get('building')
        room.capacity = request.POST.get('capacity')
        room.room_type = request.POST.get('room_type')
        room.save()
        messages.success(request, 'Classroom updated successfully')
        return redirect('classroom_list')
    return render(request, 'classroom_form.html', {'classroom': room})

@login_required
def classroom_delete(request, pk):
    room = get_object_or_404(Classroom, pk=pk)
    room.delete()
    messages.success(request, 'Classroom deleted successfully')
    return redirect('classroom_list')

@login_required
def course_list(request):
    courses = Course.objects.select_related('department', 'assigned_faculty').all()
    return render(request, 'course_list.html', {'courses': courses})

@login_required
def course_add(request):
    departments = Department.objects.all()
    faculty = Faculty.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        credit_hours = request.POST.get('credit_hours')
        course_type = request.POST.get('course_type')
        department_id = request.POST.get('department')
        faculty_id = request.POST.get('assigned_faculty')
        
        department = Department.objects.get(pk=department_id)
        course = Course.objects.create(
            name=name, code=code, credit_hours=credit_hours,
            course_type=course_type, department=department
        )
        if faculty_id:
            course.assigned_faculty = Faculty.objects.get(pk=faculty_id)
            course.save()
        messages.success(request, 'Course added successfully')
        return redirect('course_list')
    return render(request, 'course_form.html', {'departments': departments, 'faculty': faculty})

@login_required
def course_edit(request, pk):
    course = get_object_or_404(Course, pk=pk)
    departments = Department.objects.all()
    faculty = Faculty.objects.all()
    if request.method == 'POST':
        course.name = request.POST.get('name')
        course.code = request.POST.get('code')
        course.credit_hours = request.POST.get('credit_hours')
        course.course_type = request.POST.get('course_type')
        course.department = Department.objects.get(pk=request.POST.get('department'))
        faculty_id = request.POST.get('assigned_faculty')
        course.assigned_faculty = Faculty.objects.get(pk=faculty_id) if faculty_id else None
        course.save()
        messages.success(request, 'Course updated successfully')
        return redirect('course_list')
    return render(request, 'course_form.html', {'course': course, 'departments': departments, 'faculty': faculty})

@login_required
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course.delete()
    messages.success(request, 'Course deleted successfully')
    return redirect('course_list')

@login_required
def timeslot_list(request):
    day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
    time_slots = sorted(TimeSlot.objects.all(), key=lambda x: (day_order.get(x.day, 6), x.start_time))
    return render(request, 'timeslot_list.html', {'time_slots': time_slots})

@login_required
def timeslot_add(request):
    if request.method == 'POST':
        day = request.POST.get('day')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        is_break = request.POST.get('is_break') == 'on'
        TimeSlot.objects.create(day=day, start_time=start_time, 
                               end_time=end_time, is_break=is_break)
        messages.success(request, 'Time slot added successfully')
        return redirect('timeslot_list')
    return render(request, 'timeslot_form.html')

@login_required
def timeslot_edit(request, pk):
    slot = get_object_or_404(TimeSlot, pk=pk)
    if request.method == 'POST':
        slot.day = request.POST.get('day')
        slot.start_time = request.POST.get('start_time')
        slot.end_time = request.POST.get('end_time')
        slot.is_break = request.POST.get('is_break') == 'on'
        slot.save()
        messages.success(request, 'Time slot updated successfully')
        return redirect('timeslot_list')
    return render(request, 'timeslot_form.html', {'timeslot': slot})

@login_required
def timeslot_delete(request, pk):
    slot = get_object_or_404(TimeSlot, pk=pk)
    slot.delete()
    messages.success(request, 'Time slot deleted successfully')
    return redirect('timeslot_list')

@login_required
def studentgroup_list(request):
    groups = StudentGroup.objects.select_related('department').all()
    return render(request, 'studentgroup_list.html', {'groups': groups})

@login_required
def studentgroup_add(request):
    departments = Department.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        department_id = request.POST.get('department')
        year = request.POST.get('year')
        semester = request.POST.get('semester')
        semester_type = request.POST.get('semester_type')
        
        department = Department.objects.get(pk=department_id)
        StudentGroup.objects.create(
            name=name, department=department,
            year=year, semester=semester, semester_type=semester_type
        )
        messages.success(request, 'Student group added successfully')
        return redirect('studentgroup_list')
    return render(request, 'studentgroup_form.html', {'departments': departments})

@login_required
def studentgroup_edit(request, pk):
    group = get_object_or_404(StudentGroup, pk=pk)
    departments = Department.objects.all()
    if request.method == 'POST':
        group.name = request.POST.get('name')
        group.department = Department.objects.get(pk=request.POST.get('department'))
        group.year = request.POST.get('year')
        group.semester = request.POST.get('semester')
        group.semester_type = request.POST.get('semester_type')
        group.save()
        messages.success(request, 'Student group updated successfully')
        return redirect('studentgroup_list')
    return render(request, 'studentgroup_form.html', {'group': group, 'departments': departments})

@login_required
def studentgroup_delete(request, pk):
    group = get_object_or_404(StudentGroup, pk=pk)
    group.delete()
    messages.success(request, 'Student group deleted successfully')
    return redirect('studentgroup_list')

@login_required
def assignment_list(request):
    assignments = CourseAssignment.objects.select_related(
        'course', 'student_group', 'faculty'
    ).all()
    return render(request, 'assignment_list.html', {'assignments': assignments})

@login_required
def assignment_add(request):
    courses = Course.objects.all()
    groups = StudentGroup.objects.all()
    faculty = Faculty.objects.all()
    if request.method == 'POST':
        course_id = request.POST.get('course')
        group_id = request.POST.get('student_group')
        faculty_id = request.POST.get('faculty')
        
        course = Course.objects.get(pk=course_id)
        group = StudentGroup.objects.get(pk=group_id)
        faculty_member = Faculty.objects.get(pk=faculty_id)
        
        CourseAssignment.objects.create(
            course=course, student_group=group, faculty=faculty_member
        )
        messages.success(request, 'Course assignment added successfully')
        return redirect('assignment_list')
    return render(request, 'assignment_form.html', 
                 {'courses': courses, 'groups': groups, 'faculty': faculty})

@login_required
def assignment_delete(request, pk):
    assignment = get_object_or_404(CourseAssignment, pk=pk)
    assignment.delete()
    messages.success(request, 'Course assignment deleted successfully')
    return redirect('assignment_list')

@login_required
def generate_timetable(request):
    if request.method == 'POST':
        generator = TimetableGenerator()
        result = generator.generate()
        if result['success']:
            messages.success(request, result['message'])
        else:
            messages.error(request, result['message'])
        return redirect('timetable_view')
    
    stats = {
        'assignments': CourseAssignment.objects.count(),
        'time_slots': TimeSlot.objects.filter(is_break=False).count(),
        'classrooms': Classroom.objects.count(),
    }
    return render(request, 'generate_timetable.html', stats)

@login_required
def timetable_view(request):
    entries = TimetableEntry.objects.select_related(
        'course_assignment__course',
        'course_assignment__student_group',
        'course_assignment__faculty',
        'classroom',
        'time_slot'
    ).all().order_by('time_slot__day', 'time_slot__start_time')
    
    day_order = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
    grouped = {}
    for entry in entries:
        day = entry.time_slot.day
        if day not in grouped:
            grouped[day] = []
        grouped[day].append(entry)
    
    sorted_grouped = {day: grouped.get(day, []) for day in day_order if day in grouped}
    
    return render(request, 'timetable_view.html', {'grouped': sorted_grouped})

@login_required
def timetable_by_faculty(request):
    faculty_id = request.GET.get('faculty')
    if faculty_id:
        faculty = Faculty.objects.get(pk=faculty_id)
        entries = TimetableEntry.objects.filter(
            course_assignment__faculty=faculty
        ).select_related(
            'course_assignment__course',
            'course_assignment__student_group',
            'classroom',
            'time_slot'
        ).all()
        
        day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
        entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
    else:
        entries = []
        faculty = None
    
    faculty_list = Faculty.objects.all()
    return render(request, 'timetable_faculty.html', 
                 {'entries': entries, 'faculty_list': faculty_list, 'selected_faculty': faculty})

@login_required
def timetable_by_classroom(request):
    room_id = request.GET.get('classroom')
    if room_id:
        room = Classroom.objects.get(pk=room_id)
        entries = TimetableEntry.objects.filter(
            classroom=room
        ).select_related(
            'course_assignment__course',
            'course_assignment__student_group',
            'course_assignment__faculty',
            'time_slot'
        ).all()
        
        day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
        entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
    else:
        entries = []
        room = None
    
    rooms = Classroom.objects.all()
    return render(request, 'timetable_classroom.html', 
                 {'entries': entries, 'rooms': rooms, 'selected_room': room})

@login_required
def timetable_by_group(request):
    group_id = request.GET.get('group')
    if group_id:
        group = StudentGroup.objects.get(pk=group_id)
        entries = TimetableEntry.objects.filter(
            course_assignment__student_group=group
        ).select_related(
            'course_assignment__course',
            'course_assignment__faculty',
            'classroom',
            'time_slot'
        ).all()
        
        slot_data = {}
        slot_list = []
        
        all_times = TimeSlot.objects.filter(is_break=False, is_lunch=False).values('start_time', 'end_time').distinct().order_by('start_time')
        for time in all_times:
            time_key = f"{time['start_time']}-{time['end_time']}"
            slot_data[time_key] = {
                'start': time['start_time'].strftime('%H:%M'),
                'end': time['end_time'].strftime('%H:%M'),
                'monday': '',
                'tuesday': '',
                'wednesday': '',
                'thursday': '',
                'friday': ''
            }
            slot_list.append({'id': time_key, 'start': time['start_time'].strftime('%H:%M'), 'end': time['end_time'].strftime('%H:%M')})
        
        for entry in entries:
            slot = entry.time_slot
            time_key = f"{slot.start_time}-{slot.end_time}"
            if time_key in slot_data:
                slot_data[time_key][slot.day] = entry.course_assignment.course.name
        
        break_times = TimeSlot.objects.filter(is_break=True).values('start_time', 'end_time').distinct().order_by('start_time')
        for time in break_times:
            time_key = f"{time['start_time']}-{time['end_time']}"
            slot_data[time_key] = {
                'start': time['start_time'].strftime('%H:%M'),
                'end': time['end_time'].strftime('%H:%M'),
                'monday': 'BREAK',
                'tuesday': 'BREAK',
                'wednesday': 'BREAK',
                'thursday': 'BREAK',
                'friday': 'BREAK'
            }
            slot_list.append({'id': time_key, 'start': time['start_time'].strftime('%H:%M'), 'end': time['end_time'].strftime('%H:%M'), 'is_break': True})
        
        slot_list.sort(key=lambda x: x['start'])
    else:
        group = None
        slot_data = {}
        slot_list = []
    
    groups = StudentGroup.objects.all()
    
    return render(request, 'timetable_group.html', {
        'groups': groups,
        'selected_group': group,
        'slot_data': slot_data,
        'slot_list': slot_list
    })

@login_required
def entry_delete(request, pk):
    entry = get_object_or_404(TimetableEntry, pk=pk)
    entry.delete()
    messages.success(request, 'Entry deleted successfully')
    return redirect('timetable_view')

@login_required
def entry_delete_all(request):
    TimetableEntry.objects.all().delete()
    messages.success(request, 'All timetable entries cleared')
    return redirect('timetable_view')

@login_required
def download_timetable_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Day', 'Time', 'Course', 'Course Code', 'Type', 'Faculty', 'Student Group', 'Room', 'Building'])
    
    entries = TimetableEntry.objects.select_related(
        'course_assignment__course',
        'course_assignment__student_group',
        'course_assignment__faculty',
        'classroom',
        'time_slot'
    ).all().order_by('time_slot__day', 'time_slot__start_time')
    
    day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
    entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
    
    for entry in entries:
        writer.writerow([
            entry.time_slot.day.capitalize(),
            f"{entry.time_slot.start_time.strftime('%H:%M')} - {entry.time_slot.end_time.strftime('%H:%M')}",
            entry.course_assignment.course.name,
            entry.course_assignment.course.code,
            entry.course_assignment.course.course_type.capitalize(),
            entry.course_assignment.faculty.name,
            entry.course_assignment.student_group.name,
            entry.classroom.name,
            entry.classroom.building
        ])
    
    return response

@login_required
def download_timetable_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, 'openpyxl library not installed. Please run: pip install openpyxl')
        return redirect('timetable_view')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Day', 'Time', 'Course', 'Code', 'Type', 'Faculty', 'Student Group', 'Room', 'Building']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    entries = TimetableEntry.objects.select_related(
        'course_assignment__course',
        'course_assignment__student_group',
        'course_assignment__faculty',
        'classroom',
        'time_slot'
    ).all()
    
    day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
    entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
    
    for row, entry in enumerate(entries, 2):
        ws.cell(row=row, column=1, value=entry.time_slot.day.capitalize()).border = border
        ws.cell(row=row, column=2, value=f"{entry.time_slot.start_time.strftime('%H:%M')} - {entry.time_slot.end_time.strftime('%H:%M')}").border = border
        ws.cell(row=row, column=3, value=entry.course_assignment.course.name).border = border
        ws.cell(row=row, column=4, value=entry.course_assignment.course.code).border = border
        ws.cell(row=row, column=5, value=entry.course_assignment.course.course_type.capitalize()).border = border
        ws.cell(row=row, column=6, value=entry.course_assignment.faculty.name).border = border
        ws.cell(row=row, column=7, value=entry.course_assignment.student_group.name).border = border
        ws.cell(row=row, column=8, value=entry.classroom.name).border = border
        ws.cell(row=row, column=9, value=entry.classroom.building).border = border
        
        for col in range(1, 10):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
    
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    ws.row_dimensions[1].height = 25
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@login_required
def download_group_timetable_csv(request, group_id):
    group = get_object_or_404(StudentGroup, pk=group_id)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{group.name}_timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Day', 'Time', 'Course', 'Course Code', 'Type', 'Faculty', 'Room', 'Building'])
    
    entries = TimetableEntry.objects.filter(
        course_assignment__student_group=group
    ).select_related(
        'course_assignment__course',
        'course_assignment__faculty',
        'classroom',
        'time_slot'
    ).all()
    
    day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
    entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
    
    for entry in entries:
        writer.writerow([
            entry.time_slot.day.capitalize(),
            f"{entry.time_slot.start_time.strftime('%H:%M')} - {entry.time_slot.end_time.strftime('%H:%M')}",
            entry.course_assignment.course.name,
            entry.course_assignment.course.code,
            entry.course_assignment.course.course_type.capitalize(),
            entry.course_assignment.faculty.name,
            entry.classroom.name,
            entry.classroom.building
        ])
    
    return response

@login_required
def download_all_groups_timetable(request):
    groups = StudentGroup.objects.all()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="all_groups_timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Student Group', 'Semester', 'Day', 'Time', 'Course', 'Course Code', 'Type', 'Faculty', 'Room', 'Building'])
    
    for group in groups:
        entries = TimetableEntry.objects.filter(
            course_assignment__student_group=group
        ).select_related(
            'course_assignment__course',
            'course_assignment__faculty',
            'classroom',
            'time_slot'
        ).all()
        
        day_order = {'monday': 1, 'tuesday': 2, 'wednesday': 3, 'thursday': 4, 'friday': 5}
        entries = sorted(entries, key=lambda x: (day_order.get(x.time_slot.day, 6), x.time_slot.start_time))
        
        for entry in entries:
            writer.writerow([
                group.name,
                f"Sem {group.semester}",
                entry.time_slot.day.capitalize(),
                f"{entry.time_slot.start_time.strftime('%H:%M')} - {entry.time_slot.end_time.strftime('%H:%M')}",
                entry.course_assignment.course.name,
                entry.course_assignment.course.code,
                entry.course_assignment.course.course_type.capitalize(),
                entry.course_assignment.faculty.name,
                entry.classroom.name,
                entry.classroom.building
            ])
    
    return response
